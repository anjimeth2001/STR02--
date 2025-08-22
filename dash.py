import streamlit as st
import pandas as pd
import re
from io import BytesIO
from openpyxl import load_workbook
from openpyxl.styles import Font, Border, Side
from datetime import datetime

st.set_page_config(page_title="Consolidate Dye Plan — Merge & Update", layout="wide")
st.markdown("Upload all datasets to merge and download your **original workbook** with the **Dye plan** sheet updated.")

# --- Session state for uploaded files ---
if 'files_uploaded' not in st.session_state:
    st.session_state.files_uploaded = {
        'consolidate_file': None,  # <-- main workbook: Consolidate Dye plan dataset
        'finishing_file': None,
        'hank_file': None,
        'gre_file': None,
        'dye_file': None,
        'wf_file': None
    }

# --- File uploaders ---
col1, col2 = st.columns(2)
with col1:
    consolidate_file = st.file_uploader("Consolidate Dye plan (main workbook)", type=["xlsx"], key="consolidate")
    finishing_file   = st.file_uploader("Finishing PPO", type=["xlsx"], key="finishing")
    hank_file        = st.file_uploader("Hank PPO", type=["xlsx"], key="hank")
with col2:
    gre_file         = st.file_uploader("GRE Status", type=["xlsx"], key="gre")
    dye_file         = st.file_uploader("Dye PPO", type=["xlsx"], key="dye")
    wf_file          = st.file_uploader("WF PPO", type=["xlsx"], key="wf")

# --- Save uploaded files in session state ---
for key, file in zip(
    st.session_state.files_uploaded.keys(),
    [consolidate_file, finishing_file, hank_file, gre_file, dye_file, wf_file]
):
    if file is not None:
        st.session_state.files_uploaded[key] = file

# --- Display upload status with date & time ---
uploaded_files = []
file_names = [
    'Consolidate Dye plan (main)', 'Finishing PPO', 'Hank PPO',
    'GRE Status', 'Dye PPO', 'WF PPO'
]
files = list(st.session_state.files_uploaded.values())

for name, file in zip(file_names, files):
    if file:
        upload_time = datetime.now().strftime('%Y-%m-%d %H:%M:%S')
        uploaded_files.append(f"✅ {name} (Uploaded at {upload_time})")
    else:
        uploaded_files.append(f"❌ {name} (Not uploaded)")

st.write("**Upload Status:**")
col_status1, col_status2 = st.columns(2)
for i in range(0, len(uploaded_files), 2):
    col_status1.write(uploaded_files[i])
for i in range(1, len(uploaded_files), 2):
    if i < len(uploaded_files):
        col_status2.write(uploaded_files[i])

# Helper: find the latest "Dye plan ..." sheet dynamically
def pick_dye_plan_sheet(xlsx_bytes: bytes) -> str:
    xl = pd.ExcelFile(BytesIO(xlsx_bytes))
    dye_sheets = [s for s in xl.sheet_names if s.lower().startswith("dye plan")]
    if not dye_sheets:
        raise ValueError("No sheet starting with 'Dye plan' found in the main workbook.")
    def score(name: str) -> float:
        nums = re.findall(r'(\d+(?:\.\d+)?)', name)
        return float(nums[-1]) if nums else -1.0  # pick the largest trailing number like 8.22 / 8.23
    dye_sheets.sort(key=score)
    return dye_sheets[-1]

# --- Merge datasets if all uploaded ---
if all(files):
    with st.spinner("Merging datasets..."):
        # Read the main workbook bytes so we can reuse them for both pandas & openpyxl
        main_uploaded = st.session_state.files_uploaded['consolidate_file']
        main_bytes = main_uploaded.getvalue()

        # Dynamically pick the correct "Dye plan ..." sheet (e.g., Dye plan 8.22 / 8.23)
        dye_sheet_name = pick_dye_plan_sheet(main_bytes)

        # Load main sheet & others
        df_main      = pd.read_excel(BytesIO(main_bytes), sheet_name=dye_sheet_name)

        df_gre       = pd.read_excel(st.session_state.files_uploaded['gre_file'])
        df_finishing = pd.read_excel(st.session_state.files_uploaded['finishing_file'])
        df_dye       = pd.read_excel(st.session_state.files_uploaded['dye_file'])
        df_hank      = pd.read_excel(st.session_state.files_uploaded['hank_file'])
        df_wf        = pd.read_excel(st.session_state.files_uploaded['wf_file'])

        # Clean column names
        for df in [df_main, df_gre, df_finishing, df_dye, df_hank, df_wf]:
            df.columns = df.columns.str.strip()

        # --- Merge GRE Status ---
        # Assumes keys: main['Production order'] vs GRE['Origin order code']
        gre_merge = df_gre[['Origin order code', 'Receiving status', 'Last update DateTime Cmp/Div']].copy()
        df_main = df_main.merge(
            gre_merge,
            left_on='Production order',
            right_on='Origin order code',
            how='left'
        )
        df_main['Receiving status'] = df_main['Receiving status'].fillna('-')
        df_main['Last update DateTime Cmp/Div'] = df_main['Last update DateTime Cmp/Div'].fillna('-')
        if 'Origin order code' in df_main.columns:
            df_main.drop(columns=['Origin order code'], inplace=True)

        # --- Merge PPO tables into new columns ---
        def merge_ppo(df_base, df_ppo, out_col):
            # Assumes keys: df_base['Production order'] vs df_ppo['Prod Order']
            ppo = df_ppo[['Prod Order', 'Operation']].copy()
            df_base = df_base.merge(
                ppo,
                left_on='Production order',
                right_on='Prod Order',
                how='left'
            )
            df_base[out_col] = df_base['Operation'].fillna('-')
            df_base.drop(columns=['Operation', 'Prod Order'], inplace=True)
            return df_base

        df_main = merge_ppo(df_main, df_finishing, 'Finishing PPO')
        df_main = merge_ppo(df_main, df_dye,       'Dye PPO')
        df_main = merge_ppo(df_main, df_hank,      'Hank PPO')
        df_main = merge_ppo(df_main, df_wf,        'WF PPO')

        # Save merged df + source workbook bytes + sheet name for download step
        st.session_state.merged_df = df_main.copy()
        st.session_state.base_workbook_bytes = main_bytes
        st.session_state.dye_sheet_name = dye_sheet_name

    st.success(f"✅ Data merged successfully! (Updated sheet: **{st.session_state.dye_sheet_name}**)")

    with st.expander("Preview Merged Data"):
        st.dataframe(st.session_state.merged_df, use_container_width=True)

# --- Download: original workbook with the Dye plan sheet replaced by merged data ---
if all(k in st.session_state for k in ['merged_df', 'base_workbook_bytes', 'dye_sheet_name']):
    df_out = st.session_state.merged_df
    base_bytes = st.session_state.base_workbook_bytes
    sheet_name = st.session_state.dye_sheet_name

    # Load original workbook
    wb = load_workbook(BytesIO(base_bytes))

    # Remember original position of the sheet to preserve order
    if sheet_name not in wb.sheetnames:
        st.error(f"Sheet '{sheet_name}' not found in the workbook at download time.")
    else:
        idx = wb.sheetnames.index(sheet_name)
        # Remove the old sheet and recreate at the same index
        old_ws = wb[sheet_name]
        wb.remove(old_ws)
        ws = wb.create_sheet(title=sheet_name, index=idx)

        # Write DataFrame to that sheet (headers + values)
        # Header
        for c_idx, col_name in enumerate(df_out.columns, start=1):
            ws.cell(row=1, column=c_idx, value=col_name)
        # Rows
        for r_idx, row in enumerate(df_out.itertuples(index=False), start=2):
            for c_idx, value in enumerate(row, start=1):
                ws.cell(row=r_idx, column=c_idx, value=None if pd.isna(value) else value)

        # Apply formatting (Aptos Narrow 9 + thin borders) on the replaced sheet only
        thin_border = Border(left=Side(style='thin'), right=Side(style='thin'),
                             top=Side(style='thin'), bottom=Side(style='thin'))
        font = Font(name='Aptos Narrow', size=9, color='000000')
        for row in ws.iter_rows(min_row=1, max_row=ws.max_row, min_col=1, max_col=ws.max_column):
            for cell in row:
                cell.font = font
                cell.border = thin_border

        # Auto width columns on this sheet
        for col_cells in ws.columns:
            max_len = 0
            for cell in col_cells:
                v = "" if cell.value is None else str(cell.value)
                if len(v) > max_len:
                    max_len = len(v)
            ws.column_dimensions[col_cells[0].column_letter].width = max_len + 2

        # Save workbook to bytes for download
        final_output = BytesIO()
        wb.save(final_output)
        final_output.seek(0)

        st.download_button(
            label="📥 Download: Consolidate workbook (Dye plan sheet updated)",
            data=final_output,
            file_name="Consolidate_Dye_plan_UPDATED.xlsx",
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        )
